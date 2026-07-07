from __future__ import annotations

import argparse
import csv
import json
import logging
import re
from pathlib import Path
from typing import Any

from common import (
    DB_PATH,
    IMPORTS_DIR,
    clean_source_header,
    configure_logging,
    connect_db,
    discover_import_files,
    init_db,
    json_dumps,
    normalize_int,
    normalize_state,
    normalize_website,
    snake_case_name,
    utc_now_iso,
)


LOGGER = logging.getLogger(__name__)


class ImportErrorWithContext(RuntimeError):
    """Raised when a source file cannot be imported cleanly."""


def _header_key(header: str) -> str:
    text = clean_source_header(header).casefold()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _read_csv_rows(path: Path) -> list[list[Any]]:
    encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", newline="", encoding=encoding) as handle:
                return list(csv.reader(handle))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ImportErrorWithContext(f"Could not decode CSV file {path}: {last_error}")


def _read_xlsx_rows(path: Path) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportErrorWithContext("openpyxl is required to import XLSX files.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    workbook.close()
    return rows


def read_source_rows(path: Path) -> list[list[Any]]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return _read_csv_rows(path)
    if suffix in {".xlsx", ".xls"}:
        return _read_xlsx_rows(path)
    raise ImportErrorWithContext(f"Unsupported import file type: {path.suffix}")


def find_header_row(rows: list[list[Any]]) -> int:
    best_index = -1
    best_score = 0
    for index, row in enumerate(rows):
        keys = [_header_key(str(cell or "")) for cell in row]
        key_text = " | ".join(keys)
        score = 0
        if "agency name" in keys:
            score += 3
        if "agency type" in key_text:
            score += 2
        if "web site url" in key_text or "website" in key_text:
            score += 2
        if "nces" in key_text and "agency id" in key_text:
            score += 2
        if "total students" in key_text and "excludes ae" in key_text:
            score += 2
        if "state" in key_text:
            score += 1
        if score > best_score:
            best_score = score
            best_index = index
    if best_index < 0 or best_score < 5:
        raise ImportErrorWithContext("Could not identify the source header row.")
    return best_index


def _find_header_index(cleaned_headers: list[str], predicate) -> int | None:
    for index, header in enumerate(cleaned_headers):
        if predicate(_header_key(header)):
            return index
    return None


def map_headers(cleaned_headers: list[str]) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {}
    mapping["agency_name"] = _find_header_index(
        cleaned_headers,
        lambda key: key == "agency name" or key.endswith(" agency name"),
    )
    mapping["agency_id_nces"] = _find_header_index(
        cleaned_headers,
        lambda key: "agency id" in key and ("nces" in key or "assigned" in key),
    )
    mapping["state"] = _find_header_index(
        cleaned_headers,
        lambda key: key in {"location state abbr", "mailing state abbr", "state abbr"},
    )
    if mapping["state"] is None:
        mapping["state"] = _find_header_index(
            cleaned_headers,
            lambda key: key in {"state name", "state"} or key.endswith(" state"),
        )
    mapping["agency_type"] = _find_header_index(
        cleaned_headers,
        lambda key: "agency type" in key,
    )
    mapping["total_enrollment_excludes_ae"] = _find_header_index(
        cleaned_headers,
        lambda key: (
            ("total students" in key or "total enrollment" in key)
            and ("excludes ae" in key or "exclude ae" in key)
        ),
    )
    mapping["website"] = _find_header_index(
        cleaned_headers,
        lambda key: (
            ("web site" in key and "url" in key)
            or ("website" in key and ("url" in key or "address" in key))
            or key in {"website", "web address"}
        ),
    )
    return mapping


def _value(row: list[Any], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return " ".join(str(value or "").strip().split())


def _raw_json(original_headers: list[str], cleaned_headers: list[str], row: list[Any]) -> str:
    original: dict[str, Any] = {}
    cleaned: dict[str, Any] = {}
    safe_names: dict[str, Any] = {}
    for index, original_header in enumerate(original_headers):
        raw_value = row[index] if index < len(row) else ""
        clean_header = cleaned_headers[index] if index < len(cleaned_headers) else f"Column {index + 1}"
        original_key = str(original_header or f"Column {index + 1}")
        clean_key = clean_header or f"Column {index + 1}"
        safe_key = snake_case_name(clean_key)
        original[original_key] = raw_value
        cleaned[clean_key] = raw_value
        safe_names[safe_key] = raw_value
    return json.dumps(
        {"original_headers": original, "cleaned_headers": cleaned, "snake_case": safe_names},
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )


def choose_auto_source() -> Path:
    files = discover_import_files()
    if not files:
        raise ImportErrorWithContext(f"No CSV or XLSX files found in {IMPORTS_DIR}")
    selected = files[0]
    if len(files) > 1:
        LOGGER.info("Multiple import files found. Auto-selected newest file: %s", selected)
    return selected


def import_districts(source: Path | str | None = None, db_path: Path | str | None = None) -> dict[str, Any]:
    configure_logging()
    init_db(db_path)
    source_path = Path(source).expanduser().resolve() if source else choose_auto_source()
    if not source_path.exists():
        raise ImportErrorWithContext(f"Import source not found: {source_path}")

    rows = read_source_rows(source_path)
    header_index = find_header_row(rows)
    original_headers = [str(cell or "").strip() for cell in rows[header_index]]
    cleaned_headers = [clean_source_header(header) for header in original_headers]
    field_map = map_headers(cleaned_headers)

    required = ["agency_name", "state", "agency_type", "website"]
    missing = [name for name in required if field_map.get(name) is None]
    if missing:
        raise ImportErrorWithContext(
            f"Missing required source fields after header cleanup: {', '.join(missing)}"
        )

    inserted = 0
    updated = 0
    rows_read = 0
    searchable = 0
    skipped_missing_websites = 0
    states: set[str] = set()
    agency_types: set[str] = set()

    with connect_db(db_path) as conn:
        for source_row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(str(cell or "").strip() for cell in row):
                continue
            rows_read += 1
            agency_name = _value(row, field_map["agency_name"])
            if not agency_name:
                continue
            agency_id_nces = _value(row, field_map["agency_id_nces"])
            state = normalize_state(_value(row, field_map["state"]))
            agency_type = _value(row, field_map["agency_type"])
            if not state or not agency_type:
                LOGGER.debug("Skipping non-district/source note row %s: %s", source_row_number, agency_name)
                continue
            total_enrollment = normalize_int(_value(row, field_map["total_enrollment_excludes_ae"]))
            website = _value(row, field_map["website"])
            website_normalized, has_searchable_website = normalize_website(website)

            if state:
                states.add(state)
            if agency_type:
                agency_types.add(agency_type)
            if has_searchable_website:
                searchable += 1
            else:
                skipped_missing_websites += 1

            now = utc_now_iso()
            payload = {
                "source_file": source_path.name,
                "source_row_number": source_row_number,
                "agency_id_nces": agency_id_nces,
                "agency_name": agency_name,
                "state": state,
                "agency_type": agency_type,
                "total_enrollment_excludes_ae": total_enrollment,
                "website": website,
                "website_normalized": website_normalized,
                "has_searchable_website": has_searchable_website,
                "raw_json": _raw_json(original_headers, cleaned_headers, row),
                "updated_at": now,
            }

            existing = None
            if agency_id_nces:
                existing = conn.execute(
                    "SELECT id FROM districts WHERE agency_id_nces = ?",
                    (agency_id_nces,),
                ).fetchone()
            if existing is None:
                existing = conn.execute(
                    """
                    SELECT id FROM districts
                    WHERE source_file = ? AND source_row_number = ?
                    """,
                    (source_path.name, source_row_number),
                ).fetchone()

            if existing:
                payload["id"] = existing["id"]
                conn.execute(
                    """
                    UPDATE districts
                    SET source_file = :source_file,
                        source_row_number = :source_row_number,
                        agency_id_nces = :agency_id_nces,
                        agency_name = :agency_name,
                        state = :state,
                        agency_type = :agency_type,
                        total_enrollment_excludes_ae = :total_enrollment_excludes_ae,
                        website = :website,
                        website_normalized = :website_normalized,
                        has_searchable_website = :has_searchable_website,
                        raw_json = :raw_json,
                        updated_at = :updated_at
                    WHERE id = :id
                    """,
                    payload,
                )
                updated += 1
            else:
                payload["created_at"] = now
                conn.execute(
                    """
                    INSERT INTO districts (
                        source_file, source_row_number, agency_id_nces, agency_name,
                        state, agency_type, total_enrollment_excludes_ae, website,
                        website_normalized, has_searchable_website, raw_json,
                        created_at, updated_at
                    )
                    VALUES (
                        :source_file, :source_row_number, :agency_id_nces, :agency_name,
                        :state, :agency_type, :total_enrollment_excludes_ae, :website,
                        :website_normalized, :has_searchable_website, :raw_json,
                        :created_at, :updated_at
                    )
                    """,
                    payload,
                )
                inserted += 1
        conn.commit()

    summary = {
        "source_file": str(source_path),
        "header_row_number": header_index + 1,
        "rows_read": rows_read,
        "districts_inserted": inserted,
        "districts_updated": updated,
        "searchable_websites": searchable,
        "skipped_missing_websites": skipped_missing_websites,
        "states_found": sorted(states),
        "agency_types_found": sorted(agency_types),
        "field_mapping": {
            key: cleaned_headers[index] if index is not None else None
            for key, index in field_map.items()
        },
    }
    LOGGER.info("Import summary: %s", json_dumps(summary))
    return summary


def print_summary(summary: dict[str, Any]) -> None:
    print(f"Source file used: {summary['source_file']}")
    print(f"Header row: {summary['header_row_number']}")
    print(f"Rows read: {summary['rows_read']}")
    print(f"Districts inserted: {summary['districts_inserted']}")
    print(f"Districts updated: {summary['districts_updated']}")
    print(f"Searchable websites: {summary['searchable_websites']}")
    print(f"Skipped/missing websites: {summary['skipped_missing_websites']}")
    print(f"States found: {len(summary['states_found'])}")
    print(f"Agency types found: {len(summary['agency_types_found'])}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import NCES/ELSI district data into EdScanner.")
    parser.add_argument("--source", type=Path, help="CSV or XLSX file to import.")
    parser.add_argument("--db", type=Path, default=DB_PATH, help="SQLite database path.")
    parser.add_argument("--auto", action="store_true", help="Auto-select a CSV/XLSX file from imports/.")
    args = parser.parse_args()

    if not args.auto and not args.source:
        parser.error("Provide --source or --auto.")
    summary = import_districts(source=args.source if not args.auto else None, db_path=args.db)
    print_summary(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
